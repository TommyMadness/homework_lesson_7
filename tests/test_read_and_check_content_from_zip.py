import zipfile

from openpyxl.reader.excel import load_workbook

from tests.conftest import ZIP_PATH
from pypdf import PdfReader


def test_check_content_of_pdf_from_zip(pack_files_to_zip):
    with zipfile.ZipFile(ZIP_PATH, 'r') as zf:
        with zf.open("sample.pdf") as pdf_file:
            reader = PdfReader(pdf_file)

            content = reader.pages[0].extract_text()

            assert "Pellentesque sit amet lectus." in content

def test_check_content_of_xlsx_from_zip(pack_files_to_zip):
    with zipfile.ZipFile(ZIP_PATH, 'r') as zf:
        with zf.open("file_example_XLSX_1000.xlsx") as excel_file:
            workbook = load_workbook(excel_file)
            sheet = workbook.active

            content = sheet.cell(row=5, column=3).value

            assert "Hanner" in content


def test_check_content_of_csv_from_zip(pack_files_to_zip):
    with zipfile.ZipFile(ZIP_PATH, 'r') as csv_file:
        with csv_file.open("currency.csv") as file:
            content = file.read().decode("utf-8")

    assert "HRK,kn,Croatian kuna" in content
